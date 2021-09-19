#! usr/bin/python3
# createAnnexChart.py - Analyzes and prepares a summary Chart for the Thesis Annexes
# Made by Michael

""" 
Requirements:
- 109 compression specimens, maybe more
- Natalie can get a list of garbage files
- Needs Stress vs. Strain chart for each sample
- Take column B (L (mm) OD) and column C (A) and SampleID column A
- Split up to make 2 charts, one for AD and one for OD
- Garbage = 16, 42, 67
- Stress = Compressive Load / Area, where compressive load is column M
- Strain = Compressive Extension / Length (L) , where compressive extension is column K
- Need chart with:
    - Y axis = Stress
    - X axis = Strain
"""

import openpyxl, os, pprint
from pathlib import Path
from openpyxl.chart import ScatterChart, Reference, Series

def calc_stress(load, area):
    return load / area

def calc_strain(extension, length):
    return extension / length

def get_id_from_filepath(filepath):
    return int(str(filepath).split('spec')[1].split('_')[0])

def create_compression_chart(filepath_list, sub_sample_dict, data, chart_filepath, name):
    
    # Extract the data from the native files
    for file in filepath_list:

        print(f'Started processing file "{file.name}..."')

        # Get the sample's other data
        sample_id = get_id_from_filepath(file)
        sample_dict = sub_sample_dict[sample_id]
        area = sample_dict["area"]
        length = sample_dict["length"]

        # Prep the dictionary to store the data
        data[sample_id] = []

        # Open the workbook
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        last_row = sheet.max_row

        # Collect all the stress/strain data from the file
        for i in range(2, last_row):

            # Extract the raw data
            load = float(sheet['M' + str(i)].value)
            extension = float(sheet['K' + str(i)].value)

            # Calculate the stress and strain values
            stress = calc_stress(load, area)
            strain = calc_strain(extension, length)

            # Add the data to the master file
            data[sample_id].append([stress, strain])

        print(f'Finished processing file "{file.name}."')

    # Step 6: Add the calculated data into the new workbook
    workbook = openpyxl.load_workbook(chart_filepath)
    sheet = workbook.active

    # Chart formatting
    chart = ScatterChart(scatterStyle='smoothMarker')
    chart.x_axis.axPos = 'b'     # Rotates the label to be horizontal
    chart.title = f'{name} Samples Stress/Strain Curve'
    chart.height = 17
    chart.width = 34
    chart.legend = None

    # Chart axis formatting
    chart.x_axis.title = 'Strain (mm)'
    chart.y_axis.title = 'Stress (MPa)'

    for key, values in (sorted(data.items())):

        print(f'Started writing data for sample_id {key}...')
        
        # Find the next available columns and rows to add the data to
        last_col = sheet.max_column
        if last_col == 1:
            last_col -= 1

        stress_col = last_col + 1
        strain_col = last_col + 2
        start_row = 2

        # Add the headers for this key's data
        sheet.cell(row = 1, column = stress_col).value = f'ID({key})-Stress'
        sheet.cell(row = 1, column = strain_col).value = f'ID({key})-Strain'

        # Add the stress/strain data in for the key
        for i in range(len(values)):

            sheet.cell(row = start_row + i, column = stress_col).value = values[i][0]
            sheet.cell(row = start_row + i, column = strain_col).value = values[i][1]

        print(f'Finished writing data for sample_id {key}.')

        # Create a Series for the Chart with the new data
        stress_reference = Reference(sheet, min_col=stress_col, max_col=stress_col, min_row=2, max_row=len(values))
        strain_reference = Reference(sheet, min_col=strain_col, max_col=strain_col, min_row=2, max_row=len(values))
        series = Series(values=stress_reference, xvalues=strain_reference)
        chart.append(series)

    sheet.add_chart(chart, 'A1')
    workbook.save(chart_filepath)
    print(f'Finished creating Chart for {name} data.')

def create_bending_chart(filepath_list, data, filepath):

    # Extract the data from the native files
    for file in filepath_list:

        print(f'Started processing file "{file.name}..."')

        # Get the sample's id from the filepath
        sample_id = get_id_from_filepath(file)

        # Prep the dictionary to store the data
        data[sample_id] = []

        # Open the workbook
        workbook = openpyxl.load_workbook(file)
        sheet = workbook['Sheet1']
        last_row = sheet.max_row

        # Collect all the bending data
        for i in range(2, last_row):

            # Extract the raw data
            load = float(sheet['M' + str(i)].value)
            extension = float(sheet['K' + str(i)].value)

            # Add the data to the master file
            data[sample_id].append([load, extension])

        print(f'Finished processing file "{file.name}."')
    
    # Add the calculated data into the new workbook
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active

    # Chart formatting
    chart = ScatterChart(scatterStyle='smoothMarker')
    chart.x_axis.axPos = 'b'     # Rotates the label to be horizontal
    chart.title = 'Bending Samples'
    chart.height = 17
    chart.width = 25
    chart.legend = None

    # Chart axis formatting
    chart.x_axis.title = 'Compressive Extension (mm)'
    chart.y_axis.title = 'Compressive Load (N)'

    for key, values in (sorted(data.items())):

        print(f'Started writing data for sample_id {key}...')

        # Find the next available columns and rows to add the data to
        last_col = sheet.max_column
        if last_col == 1:
            last_col -= 1

        load_col = last_col + 1
        extension_col = last_col + 2
        start_row = 2

        # Add the headers for this key's data
        sheet.cell(row = 1, column = load_col).value = f'ID({key})-Load'
        sheet.cell(row = 1, column = extension_col).value = f'ID({key})-Extension'

        # Add the bending data in for the key
        for i in range(len(values)):

            sheet.cell(row = start_row + i, column = load_col).value = values[i][0]
            sheet.cell(row = start_row + i, column = extension_col).value = values[i][1]

        print(f'Finished writing data for sample_id {key}.')

        # Create a Series for the Chart with the new data
        load_reference = Reference(sheet, min_col=load_col, max_col=load_col, min_row=2, max_row=len(values))
        extension_reference = Reference(sheet, min_col=extension_col, max_col=extension_col, min_row=2, max_row=len(values))
        series = Series(values=load_reference, xvalues=extension_reference)
        chart.append(series)
        
    sheet.add_chart(chart, 'A1')
    workbook.save(filepath)
    print(f'Finished creating Chart for Bending Data.')


# Step 1: Extract Length and Area data from OnlyComp.xlsx for all samples and store into a dictionary
# Setup the directories to work with
base_dir = os.getcwd()

# Compression Data
compression_data_dir = base_dir + '/Compression_Data_Files/'
compression_sample_dir = compression_data_dir + 'OnlyComp.xlsx'
AD_chart_filename = 'AD_Chart.xlsx'
OD_chart_filename = 'OD_Chart.xlsx'
AD_chart_filepath = Path(base_dir + '/' + AD_chart_filename)
OD_chart_filepath = Path(base_dir + '/' + OD_chart_filename)

# Bending Data
bending_data_dir = base_dir + '/Bending/'
bending_chart_filename = 'Bending_Chart.xlsx'
bending_chart_filepath = Path(base_dir + '/' + bending_chart_filename)

# Open the OnlyCompy workbook for compression
sample_data_workbook = openpyxl.load_workbook(Path(compression_sample_dir))
sample_data_sheet = sample_data_workbook['OnlyComp']
sample_data_last_row = sample_data_sheet.max_row

# Extract the data
AD_sample_dict = {} # AD - Air-Dry
OD_sample_dict = {} # OD = Oven-Dry
for i in range(2, sample_data_last_row):
    
    # Extract the data from the workbook
    sample_id = sample_data_sheet['A' + str(i)].value
    length = float(sample_data_sheet['B' + str(i)].value)
    area = float(sample_data_sheet['C' + str(i)].value)
    moisture_content = sample_data_sheet['F' + str(i)].value

    # Add the data to the proper dictionary
    if moisture_content == 'AD':
        AD_sample_dict[sample_id] = {}
        AD_sample_dict[sample_id]["length"] = length
        AD_sample_dict[sample_id]["area"] = area
    elif moisture_content == 'OD':
        OD_sample_dict[sample_id] = {}
        OD_sample_dict[sample_id]["length"] = length
        OD_sample_dict[sample_id]["area"] = area

# print(pprint.pformat(sample_dict))

# Step 2: Create a new master data file for both AD and OD samples
if not AD_chart_filepath.is_file():

    # Create the new Workbook
    print('Generating new file for Air-Dried Data Chart...')
    AD_workbook = openpyxl.Workbook()
    AD_sheet = AD_workbook.active
    AD_workbook.save(AD_chart_filepath)

if not OD_chart_filepath.is_file():

    # Create the new Workbook
    print('Generating new file for Oven-Dried Data Chart...')
    OD_workbook = openpyxl.Workbook()
    OD_sheet = OD_workbook.active
    OD_workbook.save(OD_chart_filepath)

if not bending_chart_filepath.is_file():

    # Create the new Workbook
    print('Generating new file for Bending Data Chart...')
    bending_workbook = openpyxl.Workbook()
    bending_sheet = bending_workbook.active
    bending_workbook.save(bending_chart_filepath)

# Step 3: Get a list of compression data files that are air-dried and oven-dried
compression_data_files = list(Path(compression_data_dir).glob('*Compression.xlsx'))
AD_files = []
AD_keys = AD_sample_dict.keys()
OD_files = []
OD_keys = OD_sample_dict.keys()

for file in compression_data_files:
    sample_id = get_id_from_filepath(file)
    if sample_id in AD_keys:
        AD_files.append(file)
    elif sample_id in OD_keys:
        OD_files.append(file)

# Prep the dictionary to store all the data for each sample
AD_data = {}
OD_data = {}
bending_data = {}

# Get a list of bending data files
bending_data_files = list(Path(bending_data_dir).glob('*is_comp.xlsx'))

create_compression_chart(AD_files, AD_sample_dict, AD_data, AD_chart_filepath, 'Air-Dried')
create_compression_chart(OD_files, OD_sample_dict, OD_data, OD_chart_filepath, 'Oven-Dried')

create_bending_chart(bending_data_files, bending_data, bending_chart_filepath)
