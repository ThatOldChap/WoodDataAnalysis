#! usr/bin/python3
# analyzeWoodData.py - Analyzes and summarizes data files from a Resistograph
# Made for Natalie by Michael
#
# Future TODO:  1) Create a master data file
#               2) Backup that master data file on google drive or equivalent

import os, re, pprint, openpyxl, getpass
from pathlib import Path
from openpyxl.styles import Alignment, Font 
from openpyxl.chart import ScatterChart, Reference, Series

# Function that writes data to an excel file
def writeData2Spreadsheet(worksheet, data, numRows, startRow, horizAlign):
    """
    Writes a given dataset into an excel file
    
    :param worksheet: the spreadsheet that the supplied data is written to
    :param data: a list of lists that stores the data to write into the spreadsheet
    :param numRows: an integer value for the number of rows in the dataset
    :param startRow: an integer value for the starting row number in the spreadsheet
    :param horizAlign: a string value for the horizontal alignment type for the cells
    :returns: nothing
    :raises: none
    """
    for i in range(numRows):
        rowOffset = i + startRow
        # Write the data to each cell
        worksheet['A' + str(rowOffset)] = data[i][0]
        worksheet['B' + str(rowOffset)] = data[i][1]
        worksheet['C' + str(rowOffset)] = data[i][2]

        # Format the cell alignment
        worksheet['A' + str(rowOffset)].alignment = Alignment(horizontal=horizAlign)
        worksheet['B' + str(rowOffset)].alignment = Alignment(horizontal=horizAlign)
        worksheet['C' + str(rowOffset)].alignment = Alignment(horizontal=horizAlign)

# Function that finds the starting point of the dataset 
def findMatchIndex(data, colIndex, match, startIndex, direction):
    """
    Returns the index of a specified string given a search direction and index to start from

    :param data: a list of tuples that include the groups from a RegEx.findall search
    :param colIndex: the index of the tuple that is to be searched
    :param match: the string to find in the data
    :param startIndex: the index of the list to start the search at
    :param direction: the direction to search the list ('leading' for incrementing and 'trailing' for decrementing)
    :returns: an int that represents the index last instance of a specified search string
    :raises: none
    """
    if direction is 'leading':
        matchIndex = -1
        while data[matchIndex + 1][colIndex] == match:
            matchIndex += 1
            
    elif direction is 'trailing':
        matchIndex = startIndex + 1
        while data[matchIndex - 1][colIndex] == match:
            matchIndex -= 1

    return matchIndex

# Function that calculates the averages of each data column
def calcAvg(data, columnIndex, startIndex, endIndex):
    numSamples = 0
    sum = 0
    for rows in range(startIndex + 1, endIndex, 1):
        sum += int(data[rows][columnIndex])
        numSamples += 1

    return sum / numSamples

# Function for fixing grammar to append an 's'
def pluralSFix(num):
    if num is 1:
        return ''
    else:
        return 's'

# Working directory strings
workDir = ('/home/' + getpass.getuser() + '/Projects/Python_Learning/Wood_Data_Analysis/')
dataDir = workDir + 'Data_Files'
resultDir = workDir + 'Processed_Files'

# Working directory Paths
workPath = Path(workDir)
dataPath = Path(dataDir)
resultPath = Path(resultDir)
summaryFilePath = Path(resultDir + '/Results_Summary.xlsx')

# Creates the Results directory
if not resultPath.exists():
    os.makedirs(resultPath)
    print('Generated the directory: %s' % resultDir) 

# Create a list of the .txt data files to process
dataFileList = list(dataPath.glob('*.txt'))
dataFileListLen = len(dataFileList)

# List of result file paths excluding the results file
resultFileList = list(resultPath.glob('*.xlsx'))
if summaryFilePath.is_file():
    resultFileList.remove(summaryFilePath)
resultFileListLen = len(resultFileList)

# Initializing global variables
numNewFiles = 0
summaryData = []
    
# Define the RegEx to find the necessary data to gather in the file
dataRegex = re.compile(r'''(
    (\d{5})            # First column of % of Torque - Drilling Curve data
    ;                  # Separator
    (\d{5})            # Second column of % of Torque - Feed Curve data
    )''', re.VERBOSE)

# Loop through all the files in Data folder
for filePathIndex in range(dataFileListLen):

    # Skip over the file if a result file already exists
    oldFilename = dataFileList[filePathIndex].name
    newFilename = dataFileList[filePathIndex].stem.replace(' ', '_')
    newFilePath = Path(resultDir + '/' + newFilename + '.xlsx')
    if newFilePath.is_file():
        continue
    print('Processing file... %s' % oldFilename)
    
    dataFile = open(dataFileList[filePathIndex])
    dataFromFile = dataFile.read()
    dataFile.close()
    dataRaw = dataRegex.findall(dataFromFile)

    # Finds the index of the valid 0 string of data gathered from the file
    TWO_ZEROS = '00000;00000'
    ONE_ZERO = '00000'
    leadZeroIndex = findMatchIndex(dataRaw, 0, TWO_ZEROS, 0, 'leading')
    trailZeroIndex = findMatchIndex(dataRaw, 0, TWO_ZEROS, len(dataRaw) - 1, 'trailing')
    drillLeadZeroIndex = findMatchIndex(dataRaw, 1, ONE_ZERO, 0, 'leading')
    drillTrailZeroIndex = findMatchIndex(dataRaw, 1, ONE_ZERO, len(dataRaw) - 1, 'trailing')
    feedLeadZeroIndex = findMatchIndex(dataRaw, 2, ONE_ZERO, 0, 'leading')
    feedTrailZeroIndex = findMatchIndex(dataRaw, 2, ONE_ZERO, len(dataRaw) - 1, 'trailing')

    # Calculate the averages for both columns of data
    drillCurveAvg = int(round(calcAvg(dataRaw, 1, drillLeadZeroIndex, drillTrailZeroIndex), 0))
    feedCurveAvg = int(round(calcAvg(dataRaw, 2, feedLeadZeroIndex, feedTrailZeroIndex), 0))

    # Construct the data to be written to the excel file
    dataPrepped = []
    for rows in range(leadZeroIndex + 1, trailZeroIndex):
        dataPrepped.append([rows - leadZeroIndex - 1,  int(dataRaw[rows][1]), int(dataRaw[rows][2])])

    # Calculate the number of rows of data to write
    dataLen = len(dataPrepped)

    # Create the workbook and sheet for the data
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = newFilename
    sheet = wb[sheet.title]

    # Workheet formatting
    sheet.row_dimensions[1].height = 30
    sheet.column_dimensions['A'].width = 8
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 15

    # Column titles
    sheet['A1'] = 'Index'
    sheet['B1'] = 'Drill Curve\n(% of Torque)'
    sheet['C1'] = 'Feed Curve\n (% of Torque)'

    # Column formatting
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['C1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['A1'].font = Font(bold=True)
    sheet['B1'].font = Font(bold=True)
    sheet['C1'].font = Font(bold=True)

    # Label and add the average values at the top of the data columns
    sheet['A2'] = 'Average'
    sheet['B2'] = drillCurveAvg
    sheet['C2'] = feedCurveAvg
    
    # Column formatting
    sheet['A2'].alignment = Alignment(horizontal='center')
    sheet['B2'].alignment = Alignment(horizontal='center')
    sheet['C2'].alignment = Alignment(horizontal='center')
    sheet['B2'].font = Font(color='FF0000')
    sheet['C2'].font = Font(color='FF0000')
    
    # Write the gathered data into the new excel file
    writeData2Spreadsheet(sheet, dataPrepped, dataLen, 3, 'center')

    # Reference Ranges
    penetrationRef = Reference(sheet, min_col=1, min_row=3, max_col=1, max_row=dataLen+2)
    drillCurveRef = Reference(sheet, min_col=2, min_row=3, max_col=2, max_row=dataLen+2)
    feedCurveRef = Reference(sheet, min_col=3, min_row=3, max_col=3, max_row=dataLen+2)

    # Data series
    drillCurveSeries = Series(values = drillCurveRef, xvalues = penetrationRef, title='Drill Curve')
    feedCurveSeries = Series(values = feedCurveRef, xvalues = penetrationRef, title='Feed Curve')

    # Chart formatting
    chartObj = ScatterChart(scatterStyle='smoothMarker')
    chartObj.title = 'Resistograph Drill Curve vs. Feed Curve'
    chartObj.height = 15
    chartObj.width = 35  

    # Chart axis formatting
    chartObj.x_axis.title = 'Penetration (mm)'
    chartObj.y_axis.title = '% of Torque'
    chartObj.x_axis.delete = False
    chartObj.y_axis.delete = False
    chartObj.x_axis.axPos = 'b'     # Rotates the label to be horizontal
    chartObj.x_axis.scaling.max = dataPrepped[dataLen - 1][0]
    chartObj.x_axis.scaling.min = 0

    # Add the data series and create the chart
    chartObj.append(drillCurveSeries)
    chartObj.append(feedCurveSeries)
    sheet.add_chart(chartObj, 'E2')

    # Save the file after all edits are finished being made
    print('Generated new file... %s' % (newFilePath.name))
    wb.save(os.path.abspath(newFilePath))
    
    # Store the required data for the results summary file
    summaryData.append([newFilePath.name, drillCurveAvg, feedCurveAvg])
    numNewFiles += 1

# Calculate length of summaryData to be used next
sumDataLen = len(summaryData)

if not summaryFilePath.is_file():
    # Create a results summary workbook if it does not exist already
    print('Generated new file: %s' % (summaryFilePath.name))
    summaryWorkbook = openpyxl.Workbook()
    summarySheet = summaryWorkbook.active
    summarySheet.title = summaryFilePath.stem
    summarySheet = summaryWorkbook[summarySheet.title]

    # Worksheet formatting
    summarySheet.row_dimensions[1].height = 30
    summarySheet.column_dimensions['A'].width = 40
    summarySheet.column_dimensions['B'].width = 20
    summarySheet.column_dimensions['C'].width = 20

    # Column titles
    summarySheet['A1'] = 'Filename'
    summarySheet['B1'] = 'Average Drill Curve\n(% of Torque)'
    summarySheet['C1'] = 'Average Feed Curve\n (% of Torque)'

    # Column formatting
    summarySheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    summarySheet['B1'].alignment = Alignment(horizontal='center', vertical='center')
    summarySheet['C1'].alignment = Alignment(horizontal='center', vertical='center')
    summarySheet['A1'].font = Font(bold=True)
    summarySheet['B1'].font = Font(bold=True)
    summarySheet['C1'].font = Font(bold=True)

    # Refills summaryData with any missing data to recreate the full results summary file
    if sumDataLen < dataFileListLen:

        # Regenerate the resultFileList because it only has resultFiles for the files whose data is in summaryData
        # at this stage. This will be data regenerated from all the previously non-deleted result files
        resultFileList = list(resultPath.glob('*.xlsx'))
        
        for resultFileIndex in range(dataFileListLen):
            # Define the local variables used multiple times 
            resultFilePath = resultFileList[resultFileIndex]
            resultFilename = resultFilePath.name

            # Ignore files whose data has already been regenerated
            if str(summaryData).count(resultFilename) is 1:
                # Captures any files that were deleted along with the summary file
                continue
            
            # Load the spreadsheet
            resultWorkbook = openpyxl.load_workbook(resultFilePath)
            resultSheet = resultWorkbook[resultFilePath.stem]

            # Re-obtain the average values 
            drillCurveAvg = resultSheet['B2'].value
            feedCurveAvg = resultSheet['C2'].value

            # Load summaryData back with the missing data
            summaryData.append([resultFilename, drillCurveAvg, feedCurveAvg])
            numNewFiles += 1
            sumDataLen += 1
            
    # Populates the first set of data into the summary file
    writeData2Spreadsheet(summarySheet, summaryData, sumDataLen, 2, 'center')

    # Save the new results summary file
    summaryWorkbook.save(os.path.abspath(summaryFilePath))
    print('Saved %s result%s into %s' % (numNewFiles, pluralSFix(numNewFiles), summaryFilePath.name))

elif numNewFiles is 0:
    print('No new files processed.')
else:
    # Load the existing results summary workbook because there are new files that need to be added to the summary
    print('Checking file... %s' % summaryFilePath.name)
    summaryWorkbook = openpyxl.load_workbook(summaryFilePath)
    summarySheet = summaryWorkbook[summaryFilePath.stem]
    lastRow = summarySheet.max_row

    # Compile a list of the files that have already been processed
    processedFileList = []
    for i in range(lastRow):
        rowOffset = i + 2
        processedFileList.append(summarySheet['A' + str(rowOffset)].value)

    # Process summaryData and check if any of the files have already been processed
    for i in range(sumDataLen):
        rowOffset = lastRow + i + 1
        if processedFileList.count(summaryData[i][0]) > 0:
            # Skips overwriting or adding duplicate data if an already processed result file
            # was deleted and regenerated by the script
            numNewFiles -= 1
            continue
        else:
            # Update the summary file with the new results
            newSummaryDataRow = [summaryData[i][0], summaryData[i][1], summaryData[i][2]]
            writeData2Spreadsheet(summarySheet, newSummaryDataRow, 1, lastRow + 1, 'center')

    # Save the updates made to the summary file
    summaryWorkbook.save(os.path.abspath(summaryFilePath))
    print('Saved %s new result%s into %s' % (numNewFiles, pluralSFix(numNewFiles), summaryFilePath.name))


        

